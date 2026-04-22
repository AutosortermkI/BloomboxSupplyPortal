"""
PDF price list adapters.

Downloads PDF price lists from supplier websites and extracts tabular
product/price data using pdfplumber. Runs in the same pipeline as HTML
scrapers — download, parse, output ProductRow dicts.

Dependencies: pdfplumber (pip install pdfplumber)
"""
from __future__ import annotations

import io
import logging
import re
import tempfile
from pathlib import Path
from typing import Optional

from ..core.adapter import Adapter, register
from ..core.extractor import sniff_container

log = logging.getLogger("bloombox.pdf")

try:
    import pdfplumber
except ImportError:
    pdfplumber = None  # type: ignore
    log.warning("pdfplumber not installed — PDF adapters disabled")


class PDFAdapter(Adapter):
    """Base adapter for suppliers that publish PDF price lists.

    Subclasses define ``pdf_urls`` (direct download links) and override
    ``parse_table_row`` to handle their specific column layout.

    The base ``run()`` downloads each PDF, extracts tables with pdfplumber,
    and feeds rows through ``parse_table_row``.
    """

    requires_login = False
    prefer_tier = "curl_cffi"  # for downloading PDFs
    pdf_urls: list[str] = []
    skip_header_rows: int = 0  # rows to skip at top of each table

    # Subclasses override this
    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        """Convert a single table row into a product dict, or None to skip."""
        raise NotImplementedError

    def start_urls(self) -> list[str]:
        return self.pdf_urls

    def parse_page(self, html: str, url: str) -> list[dict]:
        """Override: 'html' here is actually raw PDF bytes passed through.

        When used with the standard pipeline, this won't work directly
        because the fetcher returns text. Instead, ``run()`` is overridden
        to handle the PDF binary download.
        """
        return []

    def run(self) -> list[dict]:
        """Download PDF(s) and extract product data."""
        if pdfplumber is None:
            log.error("pdfplumber not installed, cannot parse PDFs")
            return []

        all_rows: list[dict] = []
        seen: set[tuple[str, float]] = set()

        for url in self.pdf_urls:
            try:
                pdf_bytes = self._download_pdf(url)
                if not pdf_bytes:
                    continue
                rows = self._extract_from_pdf(pdf_bytes, url)
                for r in rows:
                    key = (r["name"].lower(), r["price"])
                    if key not in seen:
                        seen.add(key)
                        all_rows.append(r)
                log.info("%s: extracted %d products from %s",
                         self.supplier_name, len(rows), url)
            except Exception as e:
                log.error("%s: failed to process %s: %s",
                          self.supplier_name, url, e)

        return all_rows

    def _download_pdf(self, url: str) -> Optional[bytes]:
        """Download PDF using urllib (no fetcher tier needed for direct links)."""
        import urllib.request

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        }
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = resp.read()
                if len(data) < 1000:
                    log.warning("PDF too small (%d bytes): %s", len(data), url)
                    return None
                return data
        except Exception as e:
            log.error("Failed to download PDF from %s: %s", url, e)
            return None

    def _extract_from_pdf(self, pdf_bytes: bytes, source_url: str) -> list[dict]:
        """Extract product rows from PDF tables."""
        rows: list[dict] = []

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for table in tables:
                    for i, row in enumerate(table):
                        if i < self.skip_header_rows:
                            continue
                        # Clean up cells
                        cleaned = [
                            (cell or "").strip() for cell in row
                        ]
                        if not any(cleaned):
                            continue
                        result = self.parse_table_row(cleaned, page_num)
                        if result:
                            result.setdefault("url", source_url)
                            rows.append(result)

        return rows


# ---------------------------------------------------------------------------
# Price parsing helper for PDF text
# ---------------------------------------------------------------------------
_PRICE_RE = re.compile(r"^\$?\s*(\d{1,6}(?:,\d{3})*(?:\.\d{1,2})?)$")


def _parse_price_cell(text: str) -> Optional[float]:
    """Parse a price from a table cell (may or may not have $ sign)."""
    text = text.strip().replace(",", "")
    m = _PRICE_RE.match(text)
    if m:
        try:
            val = float(m.group(1))
            if 0 < val < 50_000:
                return val
        except ValueError:
            pass
    # Try bare number
    try:
        val = float(text)
        if 0 < val < 50_000:
            return val
    except ValueError:
        pass
    return None


# ===================================================================
# REGISTERED PDF ADAPTERS
# ===================================================================

@register
class ErnstSeedsAdapter(PDFAdapter):
    """Ernst Conservation Seeds — clean 3-column tables.

    Columns: Botanical Name | Common Name | Price per Lb
    ~427 seed products, updated annually.
    """
    supplier_id = 21
    supplier_name = "Ernst Conservation Seeds"
    pdf_urls = [
        "https://www.ernstseed.com/wp-content/uploads/2026/01/"
        "2026-Retail-Price-List-01.05.26.pdf",
    ]
    skip_header_rows = 0

    # Header patterns to skip
    _HEADERS = {"botanical name", "common name", "price per lb",
                "price per", "species", "variety"}

    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        if len(row) < 3:
            return None

        botanical = row[0].strip()
        common = row[1].strip()
        price_text = row[2].strip() if len(row) > 2 else ""

        # Skip headers and empty rows
        if not botanical or botanical.lower() in self._HEADERS:
            return None
        if common.lower() in self._HEADERS:
            return None

        price = _parse_price_cell(price_text)
        if price is None:
            return None

        # Build descriptive name
        name = common if common else botanical
        if botanical and common:
            name = f"{common} ({botanical})"

        return {
            "name": name[:200],
            "price": price,
            "unit": "per lb",
            "sku": botanical,
            "category": "Seeds",
            "container": "",
            "raw_text": f"pdf:{price_text}",
        }


@register
class ErnstBioengAdapter(PDFAdapter):
    """Ernst Conservation Seeds — bioengineering/shrub price sheet.

    Separate PDF with tree/shrub planting materials.
    Page 1 has species list, Page 2 has product forms with prices.
    Each species can be purchased in multiple forms (stakes, cuttings, whips, etc.)
    """
    supplier_id = 389  # Ernst duplicate entry for seeds
    supplier_name = "Ernst Conservation Seeds (Bioeng)"
    pdf_urls = [
        "https://www.ernstseed.com/wp-content/uploads/2025/06/"
        "Shrub-and-Tree-Planting-Material_8.5x11-Price-Sheet_"
        "FINAL_Updated-062625.pdf",
    ]
    skip_header_rows = 0

    _HEADERS = {"botanical name", "common name", "price", "species",
                "description", "size", "type"}

    def run(self) -> list[dict]:
        """Download PDF and extract species + forms with prices."""
        if pdfplumber is None:
            log.error("pdfplumber not installed, cannot parse PDFs")
            return []

        all_rows: list[dict] = []
        seen: set[tuple[str, float]] = set()

        for url in self.pdf_urls:
            try:
                pdf_bytes = self._download_pdf(url)
                if not pdf_bytes:
                    continue

                # Extract species and form prices from PDF
                rows = self._extract_bioeng_products(pdf_bytes, url)
                for r in rows:
                    key = (r["name"].lower(), r["price"])
                    if key not in seen:
                        seen.add(key)
                        all_rows.append(r)
                log.info("%s: extracted %d products from %s",
                         self.supplier_name, len(rows), url)
            except Exception as e:
                log.error("%s: failed to process %s: %s",
                          self.supplier_name, url, e)

        return all_rows

    def _extract_bioeng_products(self, pdf_bytes: bytes, source_url: str) -> list[dict]:
        """Extract species from table and form prices from text, combine them."""
        rows: list[dict] = []

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            # Extract species from page 1 table
            species_list = []
            if len(pdf.pages) > 0:
                page = pdf.pages[0]
                tables = page.extract_tables()
                if tables and len(tables) > 1:
                    table = tables[1]  # Main species table is 2nd table
                    for i, row in enumerate(table):
                        if i == 0:  # Skip header
                            continue
                        if len(row) >= 2:
                            botanical = (row[0] or "").strip()
                            common = (row[1] or "").strip()
                            if botanical and botanical.lower() not in self._HEADERS:
                                species_list.append({
                                    "botanical": botanical,
                                    "common": common,
                                })

            # Extract product forms and prices from page 2 text
            forms_with_prices = []
            if len(pdf.pages) > 1:
                page = pdf.pages[1]
                text = page.extract_text()
                if text:
                    forms_with_prices = self._parse_form_prices(text)

            # Combine: each species × each form = a product
            for species in species_list:
                for form_info in forms_with_prices:
                    name = f"{species['common']} ({species['botanical']}) — {form_info['form']}"
                    rows.append({
                        "name": name[:200],
                        "price": form_info['price'],
                        "unit": form_info.get('unit', 'per unit'),
                        "category": "Trees/Shrubs",
                        "container": form_info.get('form', ''),
                        "sku": species['botanical'],
                        "raw_text": f"pdf:form={form_info['form']},price={form_info['price']}",
                        "url": source_url,
                    })

        return rows

    def _parse_form_prices(self, text: str) -> list[dict]:
        """Parse product forms and their prices from PDF text."""
        forms = []

        # Pattern: "description....................$price unit"
        # Match lines with prices
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if not line or '$' not in line:
                continue

            # Try to parse price from line
            parts = line.rsplit('$', 1)
            if len(parts) != 2:
                continue

            form_desc = parts[0].strip()
            price_unit = parts[1].strip()

            # Skip header lines and metadata
            desc_lower = form_desc.lower()
            if any(skip in desc_lower for skip in {
                "live stakes", "unrooted cuttings", "live whips",
                "wattles & fascines", "brush or branch layering",
                "erosion control logs", "brushwood", "fees", "custom orders",
                "assessing", "establishment", "detailed ordering"
            }):
                continue

            # Clean up excessive dots from description
            form_desc = re.sub(r'\.{4,}', '', form_desc).strip()
            if not form_desc:
                continue

            # Parse price and unit from "1.50 each" or "5.00 per lin ft" etc
            price_match = re.match(r'^\$?(\d+(?:\.\d{2})?)\s+(.*)', price_unit)
            if not price_match:
                continue

            try:
                price = float(price_match.group(1))
                unit = price_match.group(2).strip()

                # Basic validation
                if 0 < price < 50_000:
                    forms.append({
                        "form": form_desc[:100],
                        "price": price,
                        "unit": unit,
                    })
            except ValueError:
                continue

        return forms

    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        """Not used — we override run() for custom extraction."""
        return None


@register
class NoltsSuppliesAdapter(PDFAdapter):
    """Nolts Greenhouse Supplies — hard goods catalog.

    52-page catalog of greenhouse supplies, pots, trays, etc.
    Table structure varies; uses text-based extraction as fallback.
    """
    supplier_id = 201
    supplier_name = "Nolts Greenhouse Supplies"
    pdf_urls = [
        "https://noltsgreenhousesupplies.com/NGScatalog.pdf",
    ]
    skip_header_rows = 0

    _SKIP = {"index", "quick", "find", "page", "nolt", "welcome",
             "our mission", "early order"}

    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        if len(row) < 2:
            return None

        # Try to find product name and price in any column arrangement
        name_parts = []
        price = None

        for cell in row:
            cell = cell.strip()
            if not cell or len(cell) < 2:
                continue
            if any(s in cell.lower() for s in self._SKIP):
                return None
            p = _parse_price_cell(cell)
            if p is not None and price is None:
                price = p
            elif len(cell) > 3:
                name_parts.append(cell)

        if not name_parts or price is None:
            return None

        name = " ".join(name_parts[:2])
        return {
            "name": name[:200],
            "price": price,
            "category": "Hard Goods",
            "container": sniff_container(name),
            "raw_text": f"pdf:{price}",
        }


@register
class BlueSkyAvailAdapter(Adapter):
    """Blue Sky Nursery — XLSX availability spreadsheet.

    They publish a weekly-updated XLSX with 2600+ product rows.
    Columns: SKU, Plant Name, On Hand, Available, Container Size, ...
    Note: No prices in this spreadsheet — availability data only.
    We still capture it as product presence data.
    """
    supplier_id = 244
    supplier_name = "Blue Sky Nursery"
    requires_login = False
    prefer_tier = "curl_cffi"

    xlsx_url = (
        "https://blueskynursery.ca/wp-content/uploads/2026/04/"
        "ONLINE-AVAILABILITY-APRIL-7-2026.xlsx"
    )

    def start_urls(self) -> list[str]:
        return [self.xlsx_url]

    def run(self) -> list[dict]:
        """Download XLSX and extract availability data."""
        try:
            import openpyxl
        except ImportError:
            log.error("openpyxl not installed, cannot parse XLSX")
            return []

        import urllib.request

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36"
            ),
        }
        try:
            req = urllib.request.Request(self.xlsx_url, headers=headers)
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = resp.read()
        except Exception as e:
            log.error("Failed to download Blue Sky XLSX: %s", e)
            return []

        rows: list[dict] = []
        seen: set[str] = set()

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb["Product"] if "Product" in wb.sheetnames else wb.active

        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 5:
                continue
            sku = str(row[0] or "").strip()
            name = str(row[1] or "").strip()
            container = str(row[4] or "").strip()

            if not name or not sku or len(name) < 3:
                continue
            if sku.lower() == "sku" or name.lower() == "name":
                continue

            # Available qty (column C or D)
            try:
                available = int(row[3] or 0)
            except (ValueError, TypeError):
                available = 0

            key = f"{sku}:{name}"
            if key in seen:
                continue
            seen.add(key)

            rows.append({
                "name": f"{name} ({container})" if container else name,
                "price": 0.0,  # No prices in availability sheet
                "sku": sku,
                "container": container,
                "in_stock": available > 0,
                "category": "Plants",
                "raw_text": f"xlsx:avail={available}",
                "extras": {"available_qty": available},
            })

        wb.close()
        log.info("Blue Sky: extracted %d products from XLSX", len(rows))
        return rows


@register
class HoffmannNurseryAdapter(PDFAdapter):
    """Hoffman Nursery (SID 111) — wholesale ornamental grasses and perennials.

    Located in Rougemont, NC. Specializes in grasses and natives.
    Note: Website is Cloudflare-protected; PDF URLs not yet publicly accessible.
    Placeholder URLs for when they publish price lists.
    """
    supplier_id = 111
    supplier_name = "Hoffman Nursery"
    pdf_urls = [
        # Placeholder URLs for when supplier publishes price list
        # "https://hoffmannursery.com/price-list.pdf",
        # "https://hoffmannursery.com/catalog/ornamental-grasses.pdf",
    ]
    skip_header_rows = 1

    _HEADERS = {"item", "description", "price", "sku", "pot size",
                "quantity", "availability", "species", "common name"}

    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        """Parse Hoffman's typical grass/perennial format.

        Expected columns: SKU | Common Name | Botanical Name | Pot Size | Price
        """
        if len(row) < 4:
            return None

        sku = row[0].strip() if row[0] else ""
        name = row[1].strip() if row[1] else ""
        botanical = row[2].strip() if len(row) > 2 else ""
        pot_size = row[3].strip() if len(row) > 3 else ""
        price_text = row[4].strip() if len(row) > 4 else ""

        # Skip headers
        if not name or name.lower() in self._HEADERS:
            return None
        if any(h in name.lower() for h in ["total", "page"]):
            return None

        # Try to extract price
        price = _parse_price_cell(price_text)
        if price is None:
            # Sometimes price might be in a different column
            for cell in row:
                p = _parse_price_cell(cell.strip())
                if p is not None:
                    price = p
                    break
            if price is None:
                return None

        # Build product name
        full_name = name
        if botanical:
            full_name = f"{name} ({botanical})"
        if pot_size:
            full_name = f"{full_name} — {pot_size}"

        return {
            "name": full_name[:200],
            "price": price,
            "sku": sku,
            "container": pot_size if pot_size else sniff_container(name),
            "unit": "per unit",
            "category": "Grasses/Perennials",
            "raw_text": f"pdf:price={price_text},sku={sku}",
        }


@register
class MountainSpringNurseryAdapter(PDFAdapter):
    """Mountain Spring Nursery (SID 121) — wholesale nursery in Reinholds, PA.

    Offers trees, shrubs, perennials, grasses. Website has login-protected catalog.
    Placeholder URLs for when they publish downloadable price lists.
    """
    supplier_id = 121
    supplier_name = "Mountain Spring Nursery"
    pdf_urls = [
        # Placeholder URLs for when supplier publishes downloadable list
        # "https://mountainspringnursery.com/catalog/price-list.pdf",
        # "https://mountainspringnursery.com/wp-content/uploads/price-list.xlsx",
    ]
    skip_header_rows = 1

    _HEADERS = {"product", "item", "price", "quantity", "sku",
                "size", "availability", "category", "common name", "botanical"}

    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        """Parse Mountain Spring's typical catalog format.

        Expected columns: SKU | Common Name | Botanical | Size | Price | Qty
        """
        if len(row) < 3:
            return None

        sku = row[0].strip() if row[0] else ""
        common = row[1].strip() if row[1] else ""
        size_or_botanical = row[2].strip() if len(row) > 2 else ""

        # Skip headers and metadata
        if not common or common.lower() in self._HEADERS:
            return None
        if any(s in common.lower() for s in ["page", "total", "catalog"]):
            return None

        # Find price in remaining columns
        price = None
        category = "Plants"

        for i, cell in enumerate(row[3:], start=3):
            cell = cell.strip()
            if not cell:
                continue
            p = _parse_price_cell(cell)
            if p is not None and price is None:
                price = p

        if price is None:
            return None

        # Determine category from common name
        name_lower = common.lower()
        if any(w in name_lower for w in ["tree", "maple", "oak", "birch", "elm"]):
            category = "Trees"
        elif any(w in name_lower for w in ["shrub", "bush", "lilac", "hydrangea"]):
            category = "Shrubs"
        elif any(w in name_lower for w in ["grass", "sedge", "rush"]):
            category = "Grasses"
        elif any(w in name_lower for w in ["native", "wildflower"]):
            category = "Natives"
        else:
            category = "Perennials"

        full_name = common
        if size_or_botanical and not size_or_botanical.isdigit():
            full_name = f"{common} ({size_or_botanical})"
        elif size_or_botanical:
            full_name = f"{common} — {size_or_botanical}"

        return {
            "name": full_name[:200],
            "price": price,
            "sku": sku,
            "container": sniff_container(common),
            "category": category,
            "unit": "per unit",
            "raw_text": f"pdf:price={price}",
        }


@register
class GoNativeTreesAdapter(PDFAdapter):
    """Go Native Trees (SID 130) — native tree specialist in Manheim, PA.

    Focuses on native eastern trees and shrubs. Has WordPress site with
    download manager but price list not publicly accessible via direct URL yet.
    """
    supplier_id = 130
    supplier_name = "Go Native Trees"
    pdf_urls = [
        # Placeholder: they use WordPress Download Manager
        # Actual URL pattern would be: https://www.gonativetrees.com/?wpdmdl=ID
        # "https://www.gonativetrees.com/price-list-native-trees.pdf",
    ]
    skip_header_rows = 1

    _HEADERS = {"species", "common name", "botanical", "price", "size",
                "container", "availability", "quantity", "notes"}

    def parse_table_row(self, row: list[str], page_num: int) -> Optional[dict]:
        """Parse Go Native Trees' native species price list format.

        Expected columns: Common Name | Botanical Name | Size | Container | Price
        They focus on native eastern species, so parse accordingly.
        """
        if len(row) < 3:
            return None

        common = row[0].strip() if row[0] else ""
        botanical = row[1].strip() if row[1] else ""
        size_text = row[2].strip() if len(row) > 2 else ""

        # Skip headers and metadata
        if not common or common.lower() in self._HEADERS:
            return None
        if any(s in common.lower() for s in ["total", "page", "note", "disclaimer"]):
            return None

        # Find price in remaining columns
        price = None
        for cell in row[3:]:
            cell = cell.strip()
            if not cell:
                continue
            p = _parse_price_cell(cell)
            if p is not None and price is None:
                price = p

        if price is None:
            return None

        # Build name with botanical info
        full_name = common
        if botanical:
            full_name = f"{common} ({botanical})"
        if size_text and not _parse_price_cell(size_text):  # Size is not a price
            full_name = f"{full_name} — {size_text}"

        container = sniff_container(common)
        if size_text and any(s in size_text for s in ["gallon", "qt", "pot", "#"]):
            container = size_text

        return {
            "name": full_name[:200],
            "price": price,
            "botanical": botanical,
            "container": container,
            "category": "Native Trees/Shrubs",
            "unit": "per unit",
            "raw_text": f"pdf:native,botanical={botanical}",
        }
